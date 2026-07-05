"""Table-aware chunker for spreadsheets (.xlsx) — the document chunker's heading-tree model does
NOT fit a grid, so this is a SEPARATE path that still emits the same `Chunk` schema (plugs into
the same ingest->chunk->embed pipeline).

Strategy A: retrieval unit = a row-group that carries its column HEADER as context, rendered as
markdown so "East region Q1 sales" matches. Per sheet: split into bands on blank rows, detect the
real header (skipping title/unit rows; carrying it across continuation bands — review T2/T3),
group data rows to a token budget, column-split wide tables so no cell is dropped. openpyxl required."""
from __future__ import annotations

import copy
import datetime
import re
from pathlib import Path

from .core import RESTRICTED_ACL, est_tokens
from .types import Chunk, ChunkResult, Element, Section

TABLE_BUDGET = (80, 600, 1200)     # (min, target, max) tokens per row-group chunk
MAX_COLS = 40                       # very wide tables (e.g. 16384 cols) -> column-split, not drop


def _cell(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, float) and v.is_integer() and abs(v) < 1e15:
        return str(int(v))
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:19]
    # markdown-safe but VALUE-PRESERVING: escape '|' (not replace), collapse newlines (review T1)
    return re.sub(r"[\r\n]+", " ", str(v)).replace("|", r"\|").strip()


def _row_cols(row):
    """index after the last non-empty cell (trailing-None-trimmed width)."""
    last = -1
    for i, v in enumerate(row):
        if v is not None and str(v).strip():
            last = i
    return last + 1


def _ncells(row):
    return sum(1 for v in row if v is not None and str(v).strip())


def _is_year(v):
    """A 4-digit year (1900-2100) is a column LABEL, not a measured value — so a row like
    ['Race', 2010, 2020] must read as a HEADER, not be misjudged as data (red-team structure-misdetect)."""
    return (isinstance(v, (int, float)) and not isinstance(v, bool)
            and float(v).is_integer() and 1900 <= v <= 2100)


def _numeric_frac(row, ncols):
    vals = [row[c] for c in range(min(ncols, len(row)))
            if row[c] is not None and str(row[c]).strip()]
    if not vals:
        return 1.0
    num = sum(1 for v in vals if isinstance(v, (int, float)) and not isinstance(v, bool)
              and not _is_year(v))                          # years are labels, not numeric data
    return num / len(vals)


def _is_header_like(row, ncols):
    """Column-name row = mostly text labels, not numbers."""
    return _numeric_frac(row, ncols) < 0.34 and _ncells(row) >= 1


def _ffill(row, ncols):
    out, last = [], ""
    for c in range(ncols):
        s = _cell(row[c]) if c < len(row) else ""
        if s:
            last = s
        out.append(last)
    return out


def _bands(rows, merges=()):
    """Split a sheet's rows into bands separated by blank rows. A row blank ONLY because it sits inside
    a VERTICAL merge (the non-top rows of a merged span are null in values_only) is NOT a separator —
    else a multi-row merged header gets shredded into header_only fragments with blank column names
    (red-team hvs_vacancy: a 5-row header split apart, the data columns left unlabeled)."""
    inside = set()
    for (r1, c1, r2, c2) in merges:
        if r2 > r1:
            inside.update(range(r1 + 1, r2 + 1))
    cur, start = [], 0
    for i, row in enumerate(rows):
        if _row_cols(row) == 0 and i not in inside:
            if cur:
                yield start, cur
            cur, start = [], i + 1
        else:
            if not cur:
                start = i
            cur.append(row)
    if cur:
        yield start, cur


def _col_slices(ncols, max_cols, keycols=1):
    """Wide tables (> max_cols) are split into column groups, each keeping the leading key
    column(s) as row context — so no cell is dropped (conservation) and each chunk stays small."""
    if ncols <= max_cols:
        return [list(range(ncols))]
    keys = list(range(min(keycols, ncols)))
    rest = list(range(len(keys), ncols))
    width = max(1, max_cols - len(keys))
    return [keys + rest[i:i + width] for i in range(0, len(rest), width)]


def _md(header, rows, cols):
    head = "| " + " | ".join(_cell(header[c]) if c < len(header) else "" for c in cols) + " |"
    sep = "| " + " | ".join("---" for _ in cols) + " |"
    body = "\n".join("| " + " | ".join(_cell(r[c]) if c < len(r) else "" for c in cols) + " |"
                     for r in rows)
    return head + "\n" + sep + ("\n" + body if body else "")


def _a1_range(ref):
    """'B4:K4' -> (r1,c1,r2,c2) 0-based; single 'A1' -> degenerate range."""
    def rc(a1):
        m = re.match(r"([A-Z]+)(\d+)", a1)
        if not m:
            return None
        col = 0
        for ch in m.group(1):
            col = col * 26 + (ord(ch) - 64)
        return (int(m.group(2)) - 1, col - 1)
    parts = ref.split(":")
    a = rc(parts[0])
    b = rc(parts[1]) if len(parts) > 1 else a
    if not a or not b:
        return None
    return (min(a[0], b[0]), min(a[1], b[1]), max(a[0], b[0]), max(a[1], b[1]))


def _parse_merges(path):
    """{sheet_title: [(r1,c1,r2,c2)]} 0-based, parsed straight from the xlsx zip WITHOUT loading any
    cell (keeps read_only streaming fast on huge sheets). values_only nulls every merged cell except
    the top-left; this geometry lets us broadcast the label back over the multi-level header."""
    import html
    import zipfile
    out = {}
    try:
        z = zipfile.ZipFile(path)
        wbx = z.read("xl/workbook.xml").decode("utf-8", "ignore")
        sheets = []
        for m in re.findall(r"<sheet\b[^>]*?/>", wbx):
            nm = re.search(r'name="([^"]+)"', m)
            rid = re.search(r'r:id="(rId\d+)"', m)
            if nm and rid:
                sheets.append((html.unescape(nm.group(1)), rid.group(1)))
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")
        rid2t = {}
        for rm in re.findall(r"<Relationship\b[^>]*/>", rels):
            i = re.search(r'Id="(rId\d+)"', rm)
            t = re.search(r'Target="([^"]+)"', rm)
            if i and t:
                rid2t[i.group(1)] = t.group(1)
        for name, rid in sheets:
            t = rid2t.get(rid, "")
            if not t:
                continue
            t = t.lstrip("/")                       # Target may be absolute '/xl/..' (openpyxl) or relative
            p = t if t.startswith("xl/") else "xl/" + t
            try:
                sx = z.read(p).decode("utf-8", "ignore")
            except Exception:
                continue
            out[name] = [r for ref in re.findall(r'<mergeCell ref="([^"]+)"', sx)
                         if (r := _a1_range(ref))]
    except Exception:
        pass
    return out


def _broadcast(layer_rows, merges, abs_start, nb):
    """Fill header-zone merged cells (null in values_only) with their top-left label, so a >=3-level
    merged header reconstructs full per-column names ('CV / 2023Q4 / E-commerce') instead of dropping
    the leaf layer. Only touches the given header rows (abs_start..abs_start+len-1)."""
    n = len(layer_rows)
    for (r1, c1, r2, c2) in merges:
        if r2 < abs_start or r1 >= abs_start + n:
            continue
        lr1 = r1 - abs_start
        tl = (layer_rows[lr1][c1] if 0 <= lr1 < n and c1 < len(layer_rows[lr1]) else None)
        if tl is None or not str(tl).strip():
            continue
        for r in range(max(r1, abs_start), min(r2 + 1, abs_start + n)):
            row = layer_rows[r - abs_start]
            while len(row) <= min(c2, nb - 1):
                row.append(None)
            for c in range(c1, min(c2 + 1, nb)):
                if c < len(row) and (row[c] is None or not str(row[c]).strip()):
                    row[c] = tl
    return layer_rows


def _read_xlsx(path):
    """openpyxl streaming read (read_only) + zip-parsed merge geometry -> ([(title, rows)], {title: merges})."""
    from openpyxl import load_workbook
    merges = _parse_merges(path)
    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheets = [(s.title, list(s.iter_rows(values_only=True))) for s in wb.worksheets]
    wb.close()
    return sheets, merges


def _read_xls(path):
    """Legacy .xls via xlrd (round-2: openpyxl can't read .xls, so it used to be a hard 0-chunk error —
    a whole document silently unindexed). Normalize '' -> None, and xlrd's half-open merge
    (rlo,rhi,clo,chi) -> our inclusive (r1,c1,r2,c2) so the same header reconstruction applies."""
    import xlrd
    try:
        book = xlrd.open_workbook(str(path), formatting_info=True)   # formatting_info -> merged_cells
    except Exception:
        book = xlrd.open_workbook(str(path))                         # fallback: lose merges, still read data
    sheets, merges = [], {}
    for sh in book.sheets():
        rows = [[(v if v != "" else None) for v in sh.row_values(r)] for r in range(sh.nrows)]
        sheets.append((sh.name, rows))
        merges[sh.name] = [(rlo, clo, rhi - 1, chi - 1)
                           for (rlo, rhi, clo, chi) in getattr(sh, "merged_cells", [])]
    return sheets, merges


def _parse_charts(path):
    """{sheet_title: [chart_text]} — each embedded chart's SEMANTIC text (title + series names + axis
    titles) from xl/charts/chartN.xml. Method A reads only cells (openpyxl) and loses charts entirely;
    for chart-heavy books (EIA STEO: 67 charts) the data story lives in the chart title/series, so a
    query 'WTI crude oil price trend' otherwise matches nothing (red-team embedded-chart). zip-parsed,
    no rendering. Sheet association comes from the first series formula 'Sheet'!$range."""
    if Path(path).suffix.lower() != ".xlsx":
        return {}                                       # .xls charts not handled (xlrd path)
    import zipfile
    out = {}
    try:
        z = zipfile.ZipFile(path)
        for n in z.namelist():
            if not re.match(r"xl/charts/chart\d+\.xml", n):
                continue
            x = z.read(n).decode("utf-8", "ignore")
            parts = []
            # two real-world encodings: Excel caches series names as <c:tx><c:v>; openpyxl writes the
            # chart title as a bare <a:t>. Take all a:t (titles/axes/rich labels) + cached series names.
            parts += [t for t in re.findall(r"<a:t>([^<]*)</a:t>", x)
                      if t.strip() and not t.strip().replace(".", "").replace("-", "").isdigit()]
            for tx in re.findall(r"<c:tx>.*?</c:tx>", x, re.S):
                parts += [v for v in re.findall(r"<c:v>([^<]*)</c:v>", tx)
                          if v.strip() and not v.strip().replace(".", "").replace("-", "").isdigit()]
            sheet = None
            for ref in re.findall(r"<c:f>([^<]+)</c:f>", x):
                if "!" in ref:
                    sheet = ref.split("!")[0].strip().strip("'")
                    break
            parts = [p for p in dict.fromkeys(parts) if p.strip()]            # dedup, keep order
            if parts:
                out.setdefault(sheet, []).append(" | ".join(parts))
    except Exception:
        pass
    return out


class TableChunker:
    """Spreadsheet chunker. `chunk(path, doc_id, lang)` -> list[Chunk]."""

    def __init__(self, target=None, min_tokens=None, max_tokens=None, max_cols=MAX_COLS):
        self.budget = (min_tokens or TABLE_BUDGET[0], target or TABLE_BUDGET[1],
                       max_tokens or TABLE_BUDGET[2])
        self.max_cols = max_cols

    def chunk(self, path, *, doc_id, lang="en", doc_meta=None, acl=None) -> list[Chunk]:
        path = Path(path)
        ext = path.suffix.lower()
        if ext == ".xlsx":
            sheets, merges_by = _read_xlsx(path)
        elif ext == ".xls":                     # round-2: legacy .xls via xlrd (was a hard 0-chunk error)
            sheets, merges_by = _read_xls(path)
        else:
            raise ValueError(f"TableChunker supports .xlsx/.xls, got {path.suffix} ({path.name})")
        book = path.stem
        out = []
        for title, rows in sheets:
            if rows:
                out.extend(self._sheet_chunks(doc_id, lang, book, title, rows, merges_by.get(title, [])))
        for sheet_title, chart_texts in _parse_charts(path).items():        # round-2: embedded charts
            crumb = [book] + ([sheet_title] if sheet_title else [])
            for ct in chart_texts:
                out.append(Chunk(
                    chunk_id="", doc_id=doc_id, kind="chart", text=ct, content_raw=None,
                    breadcrumb=crumb, section_path=" > ".join(crumb), section_id=None,
                    section_anchor=None, page_start=0, page_end=0, source_indices=[],
                    n_tokens=round(est_tokens(ct, lang)), trust="high",
                    flags=["chart_meta"] + (["sheet:" + sheet_title] if sheet_title else []), lang=lang))
        # Stamp doc-level metadata + access policy onto every chunk — SAME contract as Chunker.chunk
        # (chunker STAMPS, ingest EXTRACTS). acl defaults FAIL-CLOSED to RESTRICTED_ACL so a spreadsheet
        # ingested without an explicit policy default-denies ALL, never accidentally public
        # (INTEGRATION.md §6 ironrule #2). deepcopy -> per-chunk override is safe (no shared mutable state).
        dm = doc_meta or {}
        ac = acl if acl else RESTRICTED_ACL
        for n, c in enumerate(out):
            c.chunk_id = f"{doc_id}#{n:04d}"
            c.doc_meta = copy.deepcopy(dm)
            c.acl = copy.deepcopy(ac)
        return out

    def chunk_result(self, path, *, doc_id, lang="en", doc_meta=None, acl=None):
        """产 (elements, ChunkResult) 接 embedder 全链路(lazy-tree review#3:xlsx 此前无 element 底座、
        small-to-big 空转)。每个 table/chart chunk 合成一个 Element(idx 连续、对齐 source_indices),每 sheet
        一个 Section,回填 chunk 的 section_id/section_anchor 指向合成结构 —— 使 assemble_big 能在 sheet 内回拼
        相邻 table chunk(被行组/列切散开的同表片段)。chunk() 仍返回 list[Chunk](兼容既有调用/test_table)。"""
        from collections import OrderedDict
        chunks = self.chunk(path, doc_id=doc_id, lang=lang, doc_meta=doc_meta, acl=acl)
        elements, by_sheet = [], OrderedDict()
        for i, c in enumerate(chunks):
            sheet = c.breadcrumb[-1] if c.breadcrumb else doc_id
            elements.append(Element(idx=i, kind=c.kind, text=c.text, page=0))
            c.source_indices = [i]                                  # 对齐合成 element(idx==位置)
            by_sheet.setdefault(sheet, []).append((i, c))
        sections = []
        for sheet, items in by_sheet.items():
            idxs = [i for i, _ in items]
            sec_id = f"{doc_id}::sheet::{sheet}"
            sections.append(Section(sec_id=sec_id, doc_id=doc_id, level=1, title=str(sheet),
                                    breadcrumb=[doc_id, str(sheet)], start_idx=min(idxs),
                                    end_idx=max(idxs) + 1, parent_sec_id=None))
            for i, c in items:                                      # 同 sheet 的 chunk 归一个 section,anchor=sheet 范围
                c.section_id = sec_id
                c.section_anchor = [min(idxs), max(idxs) + 1]
        return elements, ChunkResult(chunks=chunks, sections=sections, banners=frozenset())

    def _sheet_chunks(self, doc_id, lang, book, sheet, rows, merges):
        """Per band: find the real header (first multi-col label-like row). Rows before it
        (title/units/notes) are emitted as a text block (never dropped — review T1/boe), not used
        as the header. Header-less bands continue a carried header (numeric) or become text."""
        out, carried = [], None
        for bstart, band in _bands(rows, merges):
            nb = max((_row_cols(r) for r in band), default=0)
            if nb == 0:
                continue
            h = next((i for i, r in enumerate(band) if _ncells(r) > 1 and _is_header_like(r, nb)), None)
            if h is None:                                          # no header row in this band
                if carried and any(_ncells(r) > 1 for r in band):  # data continuation -> carried header
                    out += self._emit(doc_id, lang, book, sheet, carried[0], max(carried[1], nb),
                                      band, bstart, carried[2])    # widen to this band (no col drop)
                else:                                              # annotation / 1-col list -> text
                    out += self._text_block(doc_id, lang, book, sheet, band, bstart)
                continue
            header, data_at = band[h], h + 1
            # merge-aware multi-level header (review R2: 2-row ffill drops the leaf layer of a >=3-level
            # merged header; values_only also nulls merged cells). Extend the header zone down as far as
            # any vertical merge starting inside it reaches, broadcast labels, join columns top->bottom.
            abs_h = bstart + h
            abs_end = abs_h
            changed = True
            while changed and merges:
                changed = False
                for (r1, c1, r2, c2) in merges:
                    if abs_h <= r1 <= abs_end and r2 > abs_end:
                        abs_end = r2
                        changed = True
            nlayers = abs_end - abs_h + 1
            if nlayers > 1 and h + nlayers <= len(band):
                layers = _broadcast([list(band[h + k]) for k in range(nlayers)], merges, abs_h, nb)
                header = []
                for c in range(nb):
                    parts = []
                    for lr in layers:
                        v = _cell(lr[c]) if c < len(lr) else ""
                        if v and (not parts or parts[-1] != v):     # dedup the repeats merge-broadcast makes
                            parts.append(v)
                    header.append(re.sub(r"\s+", " ", " ".join(parts)).strip())
                data_at = h + nlayers
            elif (data_at < len(band) and _ncells(band[h]) < nb     # 2-row header (group + sub labels)
                    and _is_header_like(band[data_at], nb) and _numeric_frac(band[data_at], nb) < 0.2):
                ff = _ffill(band[h], nb)
                header = [(ff[c] + " " + _cell(band[data_at][c] if c < len(band[data_at]) else "")).strip()
                          for c in range(nb)]
                data_at += 1
            if band[:h]:                                           # title/units/notes -> own text chunk
                out += self._text_block(doc_id, lang, book, sheet, band[:h], bstart)
            title = next((_cell(v) for r in band[:h] for v in r if _cell(v)), None)
            carried = (header, nb, title)
            out += self._emit(doc_id, lang, book, sheet, header, nb, band[data_at:], bstart + data_at, title)
        return out

    def _text_block(self, doc_id, lang, book, sheet, band, bstart):
        """Non-tabular band (title notes / single-column content) -> text chunk(s), budgeted."""
        crumb = [book, sheet]
        _, target, _ = self.budget
        out, lines, lstart = [], [], bstart
        rowi = bstart

        def flush(ls, s, e):
            if not ls:
                return
            text = f"Sheet: {sheet}\n" + "\n".join(ls)
            out.append(Chunk(
                chunk_id="", doc_id=doc_id, kind="table", text=text, content_raw=text,
                breadcrumb=crumb, section_path=" > ".join(crumb), section_id=None,
                section_anchor=[s, e], page_start=0, page_end=0, source_indices=list(range(s, e)),
                n_tokens=round(est_tokens(text, lang)), trust="high",
                flags=["sheet:" + sheet, "nontabular"], lang=lang))

        cur = 0.0
        for r in band:
            cells = [_cell(v) for v in r if _cell(v)]
            line = " ".join(cells)
            lt = est_tokens(line, lang)
            if lines and cur + lt > target:
                flush(lines, lstart, rowi); lines, lstart, cur = [], rowi, 0.0
            if line:
                lines.append(line); cur += lt
            rowi += 1
        flush(lines, lstart, rowi)
        return out

    def _emit(self, doc_id, lang, book, sheet, header, ncols, data, dstart, title):
        if ncols == 0:
            return []
        slices = _col_slices(ncols, self.max_cols)
        wide = len(slices) > 1
        crumb = [book, sheet] + ([title] if title else [])
        prefix = f"Sheet: {sheet}" + (f" — {title}" if title else "") + "\n"
        _, target, _ = self.budget
        out = []

        def mk(cols, grp, gs, header_only=False):
            md = prefix + _md(header, grp, cols)
            ge = gs + (1 if header_only else len(grp))
            flags = ["sheet:" + sheet]
            if wide:
                flags.append(f"cols:{cols[0]}-{cols[-1]}")
            if header_only:
                flags.append("header_only")
            out.append(Chunk(
                chunk_id="", doc_id=doc_id, kind="table", text=md, content_raw=md,
                breadcrumb=crumb, section_path=" > ".join(crumb), section_id=None,
                section_anchor=[gs, ge], page_start=0, page_end=0,
                source_indices=list(range(gs, ge)), n_tokens=round(est_tokens(md, lang)),
                trust="high", flags=flags, lang=lang))

        for cols in slices:
            if not data:                          # header-only band
                mk(cols, [], dstart - 1, header_only=True)
                continue
            htok = est_tokens(prefix + _md(header, [], cols), lang)
            group, gstart, cur = [], dstart, htok
            for r in data:
                rt = est_tokens(_md(header, [r], cols), lang)
                if group and cur + rt > target:
                    mk(cols, group, gstart)
                    gstart += len(group); group, cur = [], htok
                group.append(r); cur += rt
            if group:
                mk(cols, group, gstart)
        return out
