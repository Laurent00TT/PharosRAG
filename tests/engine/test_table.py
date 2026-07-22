"""Unit tests for the spreadsheet table chunker (Strategy A). Builds tiny .xlsx in a temp dir."""
from pathlib import Path

import pytest


from chunker.table_chunker import TableChunker, _col_slices      # noqa: E402

openpyxl = pytest.importorskip("openpyxl")


def _write(tmp_path, rows, name="t.xlsx"):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(list(r))
    p = tmp_path / name
    wb.save(p)
    return p


def test_col_slices_wide_keeps_all_columns():
    # 95-col table, max 40 -> slices each keep key col 0; union covers every column
    sl = _col_slices(95, 40)
    covered = set()
    for s in sl:
        covered |= set(s)
    assert covered == set(range(95))            # no column dropped
    assert all(s[0] == 0 for s in sl)           # key column carried into every slice


def test_header_skips_title_row(tmp_path):
    # title row + units row + real header + data — header must be the real column names
    rows = [
        ["Table 4c. Relative Standard Errors"],          # title (single cell)
        ["[Percent]"],                                    # units
        ["NAICS code", "Industry", "Total expenditures"],  # real header
        ["113", "Forestry", "11.3"],
        ["211", "Mining", "22.1"],
    ]
    res = TableChunker().chunk(_write(tmp_path, rows), doc_id="t", lang="en")
    data = [c for c in res if "nontabular" not in c.flags]
    assert data, "expected a data table chunk"
    header_line = data[0].text.split("\n")[1]
    assert "NAICS code" in header_line and "Industry" in header_line  # real header, not 'Table 4c'
    assert "Table 4c" not in header_line
    # the title text is NOT dropped — it survives in a nontabular chunk
    assert any("Table 4c" in c.text for c in res)


def test_conservation_no_cell_dropped(tmp_path):
    rows = [
        ["region", "q1", "q2"],
        ["east", 120, 130],
        ["west", 90, 95],
        ["north", 77, 88],
    ]
    res = TableChunker().chunk(_write(tmp_path, rows), doc_id="t", lang="en")
    blob = " ".join(c.text for c in res).lower()
    for tok in ["east", "west", "north", "120", "95", "88", "region", "q1"]:
        assert tok in blob                       # every value present


def test_pipe_value_preserved_not_mutated(tmp_path):
    # author list with '|' delimiter must NOT become '/'
    rows = [["rank", "authors"], [1, "Alice | Bob | Carol"]]
    res = TableChunker().chunk(_write(tmp_path, rows), doc_id="t", lang="en")
    blob = " ".join(c.text for c in res)
    assert "Alice" in blob and "Bob" in blob and "Carol" in blob
    assert "Alice / Bob" not in blob             # not slash-mutated
    assert r"\|" in blob                         # escaped, value preserved


def test_embedded_chart_semantics_extracted(tmp_path):
    # red-team embedded-chart: Method A reads only cells and loses the chart entirely. Its title and
    # series names (the data story) must surface as a chart chunk so 'WTI price trend' can match.
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Month", "WTI crude oil price"])
    for m, p in [("Jan", 70), ("Feb", 72), ("Mar", 75)]:
        ws.append([m, p])
    chart = LineChart()
    chart.title = "Crude oil price forecast"
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    chart.add_data(data, titles_from_data=True)       # series name = 'WTI crude oil price'
    ws.add_chart(chart, "E2")
    p = tmp_path / "charted.xlsx"
    wb.save(p)
    res = TableChunker().chunk(p, doc_id="t", lang="en")
    charts = [c for c in res if c.kind == "chart"]
    assert charts, "an embedded chart's title/series must produce a chart chunk"
    blob = " ".join(c.text for c in charts).lower()
    assert "crude oil" in blob                          # chart title and/or series name captured


def test_unsupported_ext_raises(tmp_path):
    # .xlsx and .xls are both supported now (round-2 added an xlrd reader); other extensions still
    # error clearly rather than silently producing nothing.
    fake = tmp_path / "data.csv"
    fake.write_text("a,b,c\n1,2,3\n")
    with pytest.raises(ValueError, match="xlsx"):
        TableChunker().chunk(fake, doc_id="t", lang="en")


# --- document metadata + access policy: the SAME fail-closed contract as the doc-chunker path
# (INTEGRATION.md §6). A spreadsheet is a second Chunk producer; it must stamp acl/doc_meta too,
# or a sensitive sheet (e.g. payroll.xlsx) ingested without a policy would be fail-OPEN. ---

def test_acl_fail_closed_by_default(tmp_path):
    # no acl passed -> RESTRICTED, never an empty/public dict. The `unset=True` sentinel is what the
    # retrieval filter denies on; a bare {} (the dataclass default) would slip past it (fail-open).
    rows = [["Name", "Salary"], ["CEO", 9_000_000]]
    res = TableChunker().chunk(_write(tmp_path, rows), doc_id="payroll", lang="en")
    assert res, "expected at least one chunk"
    for c in res:
        assert c.acl.get("unset") is True                # default-deny marker present
        assert c.acl.get("allow") == []                  # nobody allowed until permissioned
        assert c.acl.get("visibility") == "restricted"


def test_doc_meta_and_acl_stamped(tmp_path):
    # an explicit policy/metadata from ingest is stamped onto every chunk (chunker STAMPS)
    rows = [["region", "q1", "q2"], ["east", 120, 130], ["west", 90, 95]]
    res = TableChunker().chunk(_write(tmp_path, rows), doc_id="t", lang="en",
                               doc_meta={"title": "Q1 sales", "source": "erp"},
                               acl={"allow": ["grp:finance"]})
    assert res
    for c in res:
        assert c.doc_meta.get("title") == "Q1 sales" and c.doc_meta.get("source") == "erp"
        assert c.acl.get("allow") == ["grp:finance"]


def test_acl_per_chunk_independent(tmp_path):
    # acl is DEEP-copied per chunk: tightening one chunk's policy must NOT leak to its siblings.
    # a 50-col table column-splits into multiple chunks, giving us siblings to compare.
    header = [f"c{i}" for i in range(50)]
    data = [f"v{i}" for i in range(50)]
    res = TableChunker().chunk(_write(tmp_path, [header, data]), doc_id="wide", lang="en",
                               acl={"allow": ["grp:a"]})
    assert len(res) >= 2, "wide table should column-split into >=2 chunks"
    res[0].acl["allow"].append("grp:secret")             # per-chunk (per-section) override
    assert res[1].acl["allow"] == ["grp:a"]              # must NOT bleed into siblings


# --- round-2: header reconstruction (merged-header + structure-misdetect). These two were the
# critical/high defects a red-team实跑 found on real US Census tables: Strategy A's 2-row ffill drops
# the leaf layer of a 3-level merged header, and its absolute numeric_frac<0.34 threshold misreads a
# year-number row as data. Both break the column<->value binding (the component's core selling point). ---

def _write_merged(tmp_path, name, cells, merges, sheet="Sheet1"):
    """xlsx with merged cells. cells: {(row,col 1-based): value}; merges: ['B1:E1', ...]."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for (r, c), v in cells.items():
        ws.cell(row=r, column=c, value=v)
    for m in merges:
        ws.merge_cells(m)
    p = tmp_path / name
    wb.save(p)
    return p


def test_three_layer_merged_header_keeps_all_levels(tmp_path):
    # 3-layer merged header (group / sub / leaf), like Census stat tables:
    #   row1: Region(A1:A3) | 2023(B1:E1)
    #   row2:               | Q1(B2:C2)      | Q2(D2:E2)
    #   row3:               | Total | Online | Total | Online
    #   row4+: data
    # The 3rd data column (value 30/25) means "2023 / Q1 / Online" — all three levels must survive in
    # its column header. Method A's 2-row ffill drops the leaf 'Online' (read as a data row) -> fail.
    cells = {
        (1, 1): "Region", (1, 2): "2023",
        (2, 2): "Q1", (2, 4): "Q2",
        (3, 2): "Total", (3, 3): "Online", (3, 4): "Total", (3, 5): "Online",
        (4, 1): "East", (4, 2): 100, (4, 3): 30, (4, 4): 120, (4, 5): 40,
        (5, 1): "West", (5, 2): 90, (5, 3): 25, (5, 4): 110, (5, 5): 35,
    }
    merges = ["A1:A3", "B1:E1", "B2:C2", "D2:E2"]
    res = TableChunker().chunk(_write_merged(tmp_path, "stats.xlsx", cells, merges),
                               doc_id="t", lang="en")
    data = [c for c in res if "nontabular" not in c.flags]
    assert data, "expected a data table chunk"
    htext = "\n".join(c.text for c in data)
    header_line = next((l for l in htext.split("\n") if l.startswith("|")), "")
    cols = [x.strip() for x in header_line.strip("|").split("|")]
    # the Online-under-Q1-2023 column must bind ALL THREE levels together
    bound = [c for c in cols if "2023" in c and "Q1" in c and "Online" in c]
    assert bound, f"no column header bound 2023+Q1+Online together; headers={cols}"
    # and the leaf labels must NOT have leaked into the data body as a fake row
    assert "East" in htext and "30" in htext


def test_year_number_header_row_detected(tmp_path):
    # structure-misdetect: a year header row ['Race', 2010, 2020] has numeric_frac 0.67 > 0.34, so
    # Method A judges it DATA (or the whole band non-tabular) and loses column<->value binding.
    cells = {
        (1, 1): "Race", (1, 2): 2010, (1, 3): 2020,
        (2, 1): "White", (2, 2): 60.1, (2, 3): 58.2,
        (3, 1): "Black", (3, 2): 12.6, (3, 3): 13.4,
    }
    res = TableChunker().chunk(_write_merged(tmp_path, "years.xlsx", cells, []),
                               doc_id="t", lang="en")
    data = [c for c in res if "nontabular" not in c.flags]
    assert data, "year-header table should produce a DATA chunk (not all-text)"
    header_line = next((l for l in data[0].text.split("\n") if l.startswith("|")), "")
    assert "2010" in header_line and "2020" in header_line and "Race" in header_line  # years are header
    body = data[0].text
    assert "White" in body and "60.1" in body                                          # races are data


def test_merged_header_with_internal_blank_rows(tmp_path):
    # red-team hvs_vacancy: a multi-row merged header has INTERNAL blank rows (the non-top rows of a
    # vertical merge are null in values_only). _bands must NOT shred it into header_only fragments —
    # the data column must end up labeled, not blank.
    #   A1:A3=Area(vertical) | B1:C1=2023(group) | B2:B3=Q1(vertical) C2:C3=Q2(vertical)
    #   row3 is all-None (every cell is a non-top merge row) but is NOT a band separator
    cells = {(1, 1): "Area", (1, 2): "2023", (2, 2): "Q1", (2, 3): "Q2",
             (4, 1): "East", (4, 2): 10, (4, 3): 20, (5, 1): "West", (5, 2): 11, (5, 3): 21}
    merges = ["A1:A3", "B1:C1", "B2:B3", "C2:C3"]
    res = TableChunker().chunk(_write_merged(tmp_path, "blankrows.xlsx", cells, merges),
                               doc_id="t", lang="en")
    data = [c for c in res if "nontabular" not in c.flags and "header_only" not in c.flags]
    assert data, "header with internal blank rows must still yield a labeled DATA chunk"
    header_line = next((l for l in data[0].text.split("\n") if l.startswith("|")), "")
    cols = [x.strip() for x in header_line.strip("|").split("|")]
    bound = [c for c in cols if "2023" in c and "Q1" in c]
    assert bound, f"the Q1 column lost its header (band was shredded); headers={cols}"
    assert "East" in data[0].text and "10" in data[0].text


def test_chunk_result_for_embedder(tmp_path):
    # lazy-tree review#3: xlsx 进 embedder 全链路 — chunk_result 产 (elements, ChunkResult),assemble_big 能用
    from chunker import assemble_big
    rows = [["region", "q1 sales", "q2 sales"]] + [[f"region {i}", i * 10, i * 20] for i in range(120)]
    els, res = TableChunker().chunk_result(_write(tmp_path, rows), doc_id="t", lang="en")
    assert len(els) == len(res.chunks) >= 1                      # 每 chunk 一个合成 element
    assert all(e.idx == i for i, e in enumerate(els))            # idx 密集(assemble_big 断言要求)
    assert all(res.chunks[i].source_indices == [i] for i in range(len(res.chunks)))   # 对齐合成 element
    assert res.sections and all(c.section_id for c in res.chunks)   # 每 chunk 归一个 sheet section
    hit = res.chunks[0]
    big = assemble_big(hit, res.sections_by_id(), els, target=99999, min_tokens=10, max_tokens=99999,
                       acl_index=res.acl_index())
    assert big.text and big.n_tokens >= hit.n_tokens            # assemble_big 对 xlsx 工作(回拼/产出,不空转)
